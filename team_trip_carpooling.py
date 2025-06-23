# -*- coding: utf-8 -*-
"""
This script assigns passengers to drivers in hotel and airport carpools
and outputs results to an Excel workbook. It respects existing rental-car
assignments (drivers flagged by has_rental_car=True). The "Ride to Support"
column remains blank, and no Manual Review sheet is generated.
"""

import sys
import pandas as pd
from givetochat1_6 import Parser, Person  # Use definitions from givetochat1_6.py fileciteturn1file0
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import random


def simple_assign_passengers(groups):
    """
    Assigns passengers to drivers in each carpool group using round-robin,
    ensuring a maximum of 3 people per car (1 driver + up to 2 passengers).

    Args:
        groups (dict): Mapping from group key to list of Person objects in that group.

    Returns:
        dict: Mapping from group key to a dict of drivers and their assigned passengers.
    """
    assignments = {}
    for key, people in groups.items():
        drivers = [p for p in people if p.has_rental_car]
        passengers = [p for p in people if not p.has_rental_car]
        driver_slots = {d: [] for d in drivers}
        if not drivers:
            assignments[key] = driver_slots
            continue

        max_passengers_per_driver = 2
        driver_index = 0
    
        # Loop over each passenger and try to assign them
        for passenger in passengers:
            assigned = False

            # FIRST: try to find a driver with the same `app` and an open slot
            for i in range(len(drivers)):
                idx = (driver_index + i) % len(drivers)
                d = drivers[idx]
                if d.app == passenger.app and len(driver_slots[d]) < max_passengers_per_driver:
                    # Assign passenger to this same-app driver
                    driver_slots[d].append(passenger)
                    driver_index = (idx + 1) % len(drivers)
                    assigned = True
                    break

            # SECOND (fallback): if not assigned yet, assign to the next available driver regardless of `app`
            if not assigned:
                for i in range(len(drivers)):
                    idx = (driver_index + i) % len(drivers)
                    d = drivers[idx]
                    if len(driver_slots[d]) < max_passengers_per_driver:
                        driver_slots[d].append(passenger)
                        driver_index = (idx + 1) % len(drivers)
                        assigned = True
                        break
            
            if not assigned:
                print(f"Warning: Could not assign passenger {passenger.name} in group '{key}' due to full capacity")
                
        assignments[key] = driver_slots
    return assignments


def write_carpool_excel(input_xlsx: str, output_xlsx: str):
    # 1. Parse input
    all_people, h_personal, a_personal = Parser.process_excel(input_xlsx)

    # 2. Build segment groupings (no support for team trips)
    ride_hotel   = Parser.ride_to_hotel([p for p in all_people if p not in h_personal])
    ride_airport = Parser.ride_to_airport([p for p in all_people if p not in a_personal])

    # 3. Identify initial drivers and assignments
    hotel_assignments  = simple_assign_passengers(ride_hotel)
    airport_assignments = simple_assign_passengers(ride_airport)

    # identify active drivers with at least 1 passenger
    active_drivers = set()
    for cars in hotel_assignments.values():
        for driver, pax_list in cars.items():
            if pax_list: active_drivers.add(driver)
    for cars in airport_assignments.values():
        for driver, pax_list in cars.items():
            if pax_list: active_drivers.add(driver)

    assigned = {p.name: {} for p in all_people}
    for key, cars in hotel_assignments.items():
        for driver, riders in cars.items():
            for member in riders:
                assigned[member.name]['hotel'] = driver.name
    for key, cars in airport_assignments.items():
        for driver, riders in cars.items():
            for member in riders:
                assigned[member.name]['airport'] = driver.name

    # 4. Create workbook and sheet
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = 'Assignments'
    ws1.append(['Name', 'Rental Car', 'Cab to Hotel', 'Ride to Hotel', 'Ride to Support', 'Ride to Airport', 'Rental Car Given'])

    # track driver names
    driver_names = set()

    for person in sorted(all_people, key=lambda p: p.name):
        name = person.name
        rental_flag = 'YES' if person.given_rental_car and not person.has_rental_car else ''
        # Determine Rental Car column
        rental_car = person.name if person.given_rental_car or person.has_rental_car else ''

        # Ride to Hotel column: personal overrides any rental
        if person.personal['Hotel']:
            hotel_drive = 'Personal'
        elif person.has_rental_car or person.given_rental_car:
            hotel_drive = person.name
        else:
            hotel_drive = assigned[name].get('hotel', '')

        # ride to support: empty
        support_drive = ''

        # Ride to Airport column: personal overrides any rental
        if person.personal['Airport']:
            airport_drive = 'Personal'
        elif person.has_rental_car or person.given_rental_car:
            airport_drive = person.name
        else:
            airport_drive = assigned[name].get('airport', '')

        ws1.append([name, rental_car, '', hotel_drive, support_drive, airport_drive, rental_flag])

    # color sheet
    name_to_fill = {}
    for drv in active_drivers:
        r = random.randint(100, 220)
        g = random.randint(100, 220)
        b = random.randint(100, 220)
        hex_rgb = f"{r:02X}{g:02X}{b:02X}"
        name_to_fill[drv.name] = PatternFill(start_color=hex_rgb, end_color=hex_rgb, fill_type="solid")

    max_row = ws1.max_row
    for row in range(2, max_row + 1):
        for col_idx in (2, 4, 6): # only check rental car and ride columns
            cell = ws1.cell(row=row, column=col_idx)
            if cell.value in name_to_fill:
                cell.fill = name_to_fill[cell.value]

    wb.save(output_xlsx)
    print(f"Wrote carpool assignments to {output_xlsx}")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python team_trip_carpooling.py input.xlsx output.xlsx")
        sys.exit(1)
    write_carpool_excel(sys.argv[1], sys.argv[2])
