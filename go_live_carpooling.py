import sys
import pandas as pd
from collections import defaultdict
from givetochat1_6 import Parser, Person
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import random

def assign_carpool_passengers(groups):
    """
    Assign passengers to drivers for each grouping, allowing up to 3 per car.
    Drivers are those who either have a rental car or were given one.
    Passengers are assigned to drivers (round-robin), preferring same-app matches first.
    Returns a dict mapping each group key to {driver: [passengers]} assignments.
    """
    assignments = {}
    for key, people in groups.items():
        if not people:
            assignments[key] = {}
            continue
        # Identify drivers (has or given rental car) and passengers
        drivers = [p for p in people if p.has_rental_car or p.given_rental_car]
        passengers = [p for p in people if not (p.has_rental_car or p.given_rental_car)]
        # Initialize each driver with an empty passenger list
        driver_slots = {driver: [] for driver in drivers}
        # If no drivers in this group, no one can be assigned in this grouping
        if not drivers:
            assignments[key] = driver_slots  # will be empty dict
            continue
        max_passengers = 2
        driver_index = 0  # to rotate assignments among drivers
        for passenger in passengers:
            assigned = False
            # First pass: try to assign to a driver with same app (if slot available)
            for i in range(len(drivers)):
                d = drivers[(driver_index + i) % len(drivers)]
                if d.app == passenger.app and len(driver_slots[d]) < max_passengers:
                    driver_slots[d].append(passenger)
                    # move start index to next driver for fairness
                    driver_index = ((driver_index + i) % len(drivers) + 1) % len(drivers)
                    assigned = True
                    break
            # Second pass: if not assigned yet, assign to any driver with an open slot
            if not assigned:
                for i in range(len(drivers)):
                    d = drivers[(driver_index + i) % len(drivers)]
                    if len(driver_slots[d]) < max_passengers:
                        driver_slots[d].append(passenger)
                        driver_index = ((driver_index + i) % len(drivers) + 1) % len(drivers)
                        assigned = True
                        break
            if not assigned:
                # No available slot (all drivers full)
                print(f"Warning: Could not assign passenger {passenger.name} in group '{key}' (no capacity).")
        assignments[key] = driver_slots
    return assignments

def generate_carpool_assignments(input_xlsx: str, output_xlsx: str):
    # 1. Parse input Excel to Person objects and identify personal ride flags
    all_people, hotel_personal, airport_personal = Parser.process_excel(input_xlsx)
    # Filter out personal-ride people for hotel and airport grouping
    hotel_candidates = [p for p in all_people if p not in hotel_personal]
    airport_candidates = [p for p in all_people if p not in airport_personal]
    # 2. Build initial segment groupings using Parser
    ride_hotel   = Parser.ride_to_hotel(hotel_candidates)
    ride_airport = Parser.ride_to_airport(airport_candidates)
    ride_support = Parser.ride_to_support(all_people)  # support has no "personal" exclusions
    # 3. Assign drivers and passengers within each grouping
    hotel_assignments   = assign_carpool_passengers(ride_hotel)
    airport_assignments = assign_carpool_passengers(ride_airport)
    support_assignments = assign_carpool_passengers(ride_support)
    # 4. Step 2: extend support assignments by filling open seats with remaining unassigned travelers
    # Collect all support drivers and unassigned people
    unassigned_support = []  # people without a support assignment and no rental

    # helper method for separating ID/IE/Exemplar/Emeritus from carpooling
    def special(p: Person) -> bool:
        return p.app in ("ID", "IE") or any(kw in p.name.lower() for kw in ("emeritus", "exemplar"))

    for group_key, cars in support_assignments.items():
        # If group has no drivers, all people in it are unassigned (none has rental)
        if not cars:
            for p in ride_support.get(group_key, []):
                # Only consider those who truly have no car
                if not (p.has_rental_car or p.given_rental_car):
                    unassigned_support.append(p)
        else:
            # Group has some drivers; find any passengers not assigned due to capacity
            # (The assignment function already tried to assign all it could, 
            # so any leftover in ride_support group with no driver are unassigned)
            assigned_passengers = {passenger for driver, plist in cars.items() for passenger in plist}
            for p in ride_support.get(group_key, []):
                if p not in cars.keys() and p not in assigned_passengers:
                    # p is not a driver and not assigned as passenger in this group
                    if not (p.has_rental_car or p.given_rental_car):
                        unassigned_support.append(p)
    # Sort unassigned folks by onsite duration, grouping those with longer onsite durations first
    # Folks with longer onsite durations will have less potential driver matches - they get "dibs"
    unassigned_support.sort(key=lambda p: p.end_onsite - p.begin_onsite, reverse=True)
    # Now try to assign each leftover person to any driver with matching time and available seat
    for group_key, cars in support_assignments.items():
        for driver, passengers in list(cars.items()):
            if special(driver): continue # don't add passenger to ID/IE/Exemplar/Emeritus driver
            # If driver's car is not full, attempt to find matches
            while len(passengers) < 2:
                # Find any unassigned person who can fit with this driver
                match = None
                # First look for same-app matches
                for p in unassigned_support:
                    if p.app == driver.app:
                        # Check time-of-day alignment and containment of onsite window
                        if (p.location == driver.location and 
                            p.hotel == driver.hotel and 
                            p.begin_onsite.time() == driver.begin_onsite.time() and 
                            p.end_onsite.time() == driver.end_onsite.time() and
                            p.begin_onsite >= driver.begin_onsite and 
                            p.end_onsite <= driver.end_onsite):
                            match = p
                            break
                # If none found with same app, look for any app
                if match is None:
                    for p in unassigned_support:
                        if (p.location == driver.location and 
                            p.hotel == driver.hotel and 
                            p.begin_onsite.time() == driver.begin_onsite.time() and 
                            p.end_onsite.time() == driver.end_onsite.time() and
                            p.begin_onsite >= driver.begin_onsite and 
                            p.end_onsite <= driver.end_onsite):
                            match = p
                            break
                if match is None:
                    break  # no suitable passenger for this driver
                # Assign the found passenger to this driver's car
                passengers.append(match)
                # Remove the passenger from unassigned list
                unassigned_support.remove(match)
            # update the entry (since we modified passengers list in place)
            support_assignments[group_key][driver] = passengers
    # 5. Prepare output workbook and sheet
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Assignments'
    ws.append(['Name', 'Rental Car', 'Cab to Hotel', 'Ride to Hotel', 'Ride to Support', 'Ride to Airport', 'Rental Car Given'])
    # Map each passenger name to their driver for each segment
    assigned_driver = {p.name: {'hotel': '', 'support': '', 'airport': ''} for p in all_people}
    # Record assignments for Hotel and Airport segments
    active_drivers = set() # collect drivers with 1+ passenger for coloring
    for cars in hotel_assignments.values():
        for driver, pax_list in cars.items():
            if pax_list:  # only consider drivers with passengers
                active_drivers.add(driver.name)
            for pax in pax_list:
                assigned_driver[pax.name]['hotel'] = driver.name
    for cars in airport_assignments.values():
        for driver, pax_list in cars.items():
            if pax_list:  # only consider drivers with passengers
                active_drivers.add(driver.name)
            for pax in pax_list:
                assigned_driver[pax.name]['airport'] = driver.name
    # Record assignments for Support segment
    for cars in support_assignments.values():
        for driver, pax_list in cars.items():
            if pax_list:  # only consider drivers with passengers
                active_drivers.add(driver.name)
            for pax in pax_list:
                assigned_driver[pax.name]['support'] = driver.name
    # Populate the sheet rows
    for person in sorted(all_people, key=lambda p: p.name):
        name = person.name
        rental_car = person.name if (person.has_rental_car or person.given_rental_car) else ''
        # Ride to Hotel
        if person.personal.get("Hotel", False):
            ride_hotel = 'Personal'
        elif rental_car != '':
            ride_hotel = person.name
        else:
            ride_hotel = assigned_driver[name]['hotel']
        # Ride to Support
        if rental_car != '':
            ride_support = person.name
        else:
            ride_support = assigned_driver[name]['support']
        # Ride to Airport
        if person.personal.get("Airport", False):
            ride_airport = 'Personal'
        elif rental_car != '':
            ride_airport = person.name
        else:
            ride_airport = assigned_driver[name]['airport']
        # Rental car given
        rental_given = 'YES' if (person.given_rental_car and not person.has_rental_car) else ''
        ws.append([name, rental_car, '', ride_hotel, ride_support, ride_airport, rental_given])
    # Apply color fill: each driver gets a unique color, fill any cell containing that driver's name
    color_map = {}
    for drv in active_drivers:
        # generate a pastel-ish color for visibility
        r = random.randint(100, 220)
        g = random.randint(100, 220)
        b = random.randint(100, 220)
        hex_color = f"{r:02X}{g:02X}{b:02X}"
        color_map[drv] = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    # Fill cells in Rental Car and Ride columns if value matches a driver name
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):  # columns 2-5 are rental car and ride columns
        for cell in row:
            if cell.value in color_map:
                cell.fill = color_map[cell.value]
    # Save to output file
    wb.save(output_xlsx)
    print(f"Carpool assignments written to {output_xlsx}")

# If run as a script, take input and output file paths from command-line arguments
if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python go-live_carpooling.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    generate_carpool_assignments(sys.argv[1], sys.argv[2])
